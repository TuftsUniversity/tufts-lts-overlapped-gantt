#!/usr/bin/env python3
# import pandas as pd

import pandas as pd
import plotly.express as pex
import matplotlib.pyplot as plt
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import sys
import kaleido
from datetime import datetime
from datetime import date
from itertools import cycle
from matplotlib.pyplot import cm
import numpy as np
import re
#import plotly.graph_objects as go
filename = askopenfilename(title="Select project sheet")


wb = load_workbook(filename)
ws = wb.active # or wb["SheetName"] # <- change the name here

rows = [[c.value for c in r] for r in ws.iter_rows()
         if not ws.row_dimensions[r[0].row].hidden]

projects_df = pd.DataFrame(data= rows[1:], columns=rows[0], dtype=str)
projects_df['level_of_effort'] = projects_df['level_of_effort'].astype("float")
projects_df['level_of_effort'] = projects_df['level_of_effort'].astype("Int64")


projects_df = projects_df[['task', 'level_of_effort', 'start_date', 'end_date']].copy()
projects_df['stack'] = 0
projects_df = projects_df.sort_values(['start_date', 'end_date', 'level_of_effort'])
pd.set_option('display.max_columns', None)
pd.options.display.max_colwidth = 200
print(projects_df)

projects_df = projects_df.dropna()
projects_df['start_date'] = projects_df['start_date'].apply(lambda x: x.replace(" 00:00:00", ""))
projects_df['end_date'] = projects_df['end_date'].apply(lambda x: x.replace(" 00:00:00", ""))




min_start_date = projects_df['start_date'].apply(lambda x: datetime.strptime(x, "%Y-%m-%d")).min()

max_end_date = projects_df['end_date'].apply(lambda x: datetime.strptime(x, "%Y-%m-%d")).max()#  max_end_date = projects_df.max(pd.to_datetime('finish'))
delta = max_end_date - min_start_date
length_of_matrix = delta

delta = int(delta.total_seconds()/60/60/24)
height_of_matrix = int(projects_df['level_of_effort'].sum())

print("Height of matrix: " + str(height_of_matrix))

print("Width of Matrix:  " + str(delta))
rows, cols = (delta, height_of_matrix)
arr = [[0]*cols]*rows

date_range = pd.date_range(min_start_date, max_end_date)

range_list = list(reversed(list(range(0, height_of_matrix))))

for z in range(0, len(range_list)):

    range_list[z] = str(range_list[z])

master_plotting_df = pd.DataFrame(columns=date_range, index=range_list)

master_plotting_df = master_plotting_df.applymap(lambda x: 0)



project_plotting_df = master_plotting_df.copy()
project_plotting_df = master_plotting_df.copy()
for x in range(0, len(projects_df)):

    y = 0
    while y + int(projects_df.iloc[x, projects_df.columns.get_loc('level_of_effort')]) < height_of_matrix:

        # create a block the height and width of the project level of effort and the start and end dates
        # the start and end dates are fixed, but where a project of this level of effort fits on the y axis will depend where it "nests"
        # among other projects
        project_dates_and_effort_df = project_plotting_df.loc[str(y + int(projects_df.iloc[x, projects_df.columns.get_loc('level_of_effort')]) - 1): str(y), projects_df.iloc[x, projects_df.columns.get_loc('start_date')]: projects_df.iloc[x, projects_df.columns.get_loc('end_date')]]


        master_plotting_subslice_df = master_plotting_df.copy()


        #this part of the code blocks out parts of the master dataframe if the rectangle of the x run of date range (fixed) fits into a block with the height of level of effort if
        # this starts at a given y value.     if not y is incremented and the next one up is tried

        # the master plotting df has a value of 0 in its initial state in this sector if no part of it has been occupied by other sectors
        # this the DataFrame equality check will confirm that for equal rows and columns they will all have the same value for each cell "0"
        # but if other projects have occupied part of this sector some of the them will have a value of 1 and therefore the equality operator will
        # not return that they are equivalent, dnd y will be incrermented 1 and this will be tried again
        if (project_dates_and_effort_df.equals(master_plotting_df.loc[str(y + int(projects_df.iloc[x, projects_df.columns.get_loc('level_of_effort')]) - 1): str(y), projects_df.iloc[x, projects_df.columns.get_loc('start_date')]: projects_df.iloc[x, projects_df.columns.get_loc('end_date')]])):
            projects_df.iloc[x, projects_df.columns.get_loc('stack')] = y

            master_plotting_df.loc[str(y + int(projects_df.iloc[x, projects_df.columns.get_loc('level_of_effort')]) - 1): str(y), projects_df.iloc[x, projects_df.columns.get_loc('start_date')]: projects_df.iloc[x, projects_df.columns.get_loc('end_date')]] =  master_plotting_df.loc[str(y + int(projects_df.iloc[x, projects_df.columns.get_loc('level_of_effort')]) - 1): str(y), projects_df.iloc[x, projects_df.columns.get_loc('start_date')]: projects_df.iloc[x, projects_df.columns.get_loc('end_date')]].applymap(lambda z: 1)



            y += 1
            break


        else:
            y += 1
            #print('miss')
#master_plotting_df.to_excel("Plotting DataFrame.xlsx")
print(projects_df)

new_max_height_df = projects_df.copy()

new_max_height = projects_df['stack'].max()

new_max_height_df = projects_df[projects_df['stack'] == new_max_height]

new_max_height_plus_level_of_effort = int(new_max_height_df['level_of_effort'].max()) + new_max_height
# print(new_max_height_df)
# print(new_max_height_plus_level_of_effort)




df = projects_df.copy()

fig, gnt = plt.subplots(figsize=(16, 10))
array = np.linspace(0, 1, len(df))
np.random.shuffle(array)
color = iter(cm.rainbow(array))
#
# Plotting the area chart
# palette = cycle(pex.colors.qualitative.Bold)
# plt.style.use('ggplot')

df = df.reset_index()
for l in range(0, len(df)):
    #print(df.loc[l, 'start'])
    #print(df.loc[l, 'finish'])
    start = datetime.strptime(df.loc[l, 'start_date'], "%Y-%m-%d")
    #print(pd.to_datetime(start))
    #print(type(pd.to_datetime(start)))
    finish = datetime.strptime(df.loc[l, 'end_date'], "%Y-%m-%d")
    #print(type(pd.to_datetime(finish)))
    #print(type(pd.to_datetime(finish)))
    #gnt.broken_barh([(pd.to_datetime(start).timestamp(), pd.to_datetime(finish).timestamp()-pd.to_datetime(start).timestamp())], [int(df.loc[l, 'bandwidth']), int(df.loc[l, 'effort'])], color=next(color))
    gnt.broken_barh([(pd.to_datetime(start), pd.to_datetime(finish)-pd.to_datetime(start))], [int(df.loc[l, 'stack']), int(df.loc[l, 'level_of_effort'])], color=next(color), label=df.loc[l, 'task'])

    data = [(pd.to_datetime(start), pd.to_datetime(finish)-pd.to_datetime(start))]

    # print(data)
    for x1, x2 in data:
        gnt.text(x= x1 + x2/2,
                y= (int(df.loc[l, 'stack']) + int(df.loc[l, 'level_of_effort'])) - int(df.loc[l, 'level_of_effort'])/2,
                s=df.loc[l,  'task'],
                ha='center',
                va='center',
                color='blue',
                fontsize='xx-small',

               )

# ax.set_yticks(range(len(label)))
# ax.set_yticklabels(label)


# gantt = pex.timeline(df, x_start='start', x_end='finish',
#                      y='bandwidth', color='task', height=600, width=height)
# for i, d in enumerate(gantt.data):
#     gantt.data[i]['width'] = df.loc[x, 'width']
#
# fig.set_figheight(15)
# fig.set_figwidth(80)
# gnt.set_position([0, 0, 1, .6])
fig.tight_layout()
# fig.tight_layout()
gnt.set_xlabel("Date")
gnt.set_ylabel("Hours per Day")
#gnt.update_traces(marker_color=df.loc['task'].tolist(), marker_colorscale="Rainbow")
#gnt.set_ylabel('fruit supply')
#fig.show()
# gantt.update_traces(width=width_list)
#fig.write_image("yourfile.png")

fig.legend(ncol=3)

top_value_benchmark = .710/10

top_value = top_value_benchmark * new_max_height_plus_level_of_effort

plt.subplots_adjust(left=0.1, right=0.9, bottom=0.2, top=0.9)
figsize=(16, 10)
plt.xticks(rotation=45)
# plt.update_layout(legend=dict(font=dict(size(10))))
plt.show(block=True)
