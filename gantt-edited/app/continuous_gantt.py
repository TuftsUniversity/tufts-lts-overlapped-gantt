import pandas as pd
import plotly.express as pex
import matplotlib.pyplot as plt
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from datetime import datetime
from matplotlib.pyplot import cm
import numpy as np
import os
from io import BytesIO
import logging
import textwrap
import io

from flask import Flask, request, jsonify, current_app

def wrap_text(text, width=20):
    """Wraps the text to a specified width."""
    return '\n'.join(textwrap.wrap(text, width=width))


def generate_gantt_chart(file_content):
    try:
        # Load the Excel workbook from the in-memory file content
        file_stream = BytesIO(file_content)
        wb = load_workbook(file_stream)
        ws = wb.active

        rows = [[c.value for c in r] for r in ws.iter_rows() if not ws.row_dimensions[r[0].row].hidden]

        projects_df = pd.DataFrame(data= rows[1:], columns=rows[0], dtype=str)

        
       # Check for different spellings of 'level_of_effort'
        # if 'level_of_effort' not in projects_df.columns and 'Level of Effort' not in projects_df.columns:
        #     # If neither spelling is present, create the column with a default value of 2
        #     projects_df["level_of_effort"] = 2
        # else:
            # If 'Level of Effort' is present, rename it to 'level_of_effort' for consistency
       # 'Level of Effort' in projects_df.columns:
        projects_df = projects_df.rename(columns={'Level of Effort': 'level_of_effort'})

    
        projects_df = projects_df.dropna()
        pd.options.display.max_colwidth = 200
      
        if 'level_of_effort' in projects_df.columns:
            projects_df['level_of_effort'] = (
                projects_df['level_of_effort']
                .str.replace('.0', '', regex=False)  # Remove '.0' from float-like strings
                .replace('nan', '') 
                .replace('', '')
                .astype(int)  # Convert to integer
            )
    
        
        projects_df["stack"] = 0
        
        projects_df = projects_df.rename(columns={"Start date": 'start_date', "Due date": 'end_date'})


        projects_df = projects_df.dropna(subset=["start_date", "end_date"])
        projects_df["start_date"] = projects_df["start_date"].apply(
            lambda x: x.replace(" 00:00:00", "")
        )
        projects_df["end_date"] = projects_df["end_date"].apply(
            lambda x: x.replace(" 00:00:00", "")
        )

        # Convert 'start_date' and 'end_date' columns to datetime, handling errors gracefully
        projects_df["start_date"] = pd.to_datetime(projects_df["start_date"], errors='coerce')
        projects_df["end_date"] = pd.to_datetime(projects_df["end_date"], errors='coerce')
        if (projects_df["start_date"] > projects_df["end_date"]).any():
            raise ValueError("Error: Start date cannot be later than the end date.")

            
        # Drop rows where dates couldn't be converted to datetime (NaT)
        projects_df = projects_df.dropna(subset=["start_date", "end_date"])

      # Extract the minimum and maximum dates safely
        min_start_date = projects_df["start_date"].min()
        max_end_date = projects_df["end_date"].max()

      
        delta = max_end_date - min_start_date
        length_of_matrix = delta

        delta = int(delta.total_seconds() / 60 / 60 / 24)
        height_of_matrix = int(projects_df["level_of_effort"].sum())

        rows, cols = (delta, height_of_matrix)
        arr = [[0] * cols] * rows

        date_range = pd.date_range(min_start_date, max_end_date)

        range_list = list(reversed(list(range(0, height_of_matrix))))

        for z in range(0, len(range_list)):
            range_list[z] = str(range_list[z])

        projects_df = projects_df.sort_values(
            ["start_date", "end_date", "level_of_effort"],
            ascending=[True, False, True]  # Specify the sorting order for each column
        )
        master_plotting_df = pd.DataFrame(columns=date_range, index=range_list)
        master_plotting_df = master_plotting_df.applymap(lambda x: 0)

        project_plotting_df = master_plotting_df.copy()
        for x in range(0, len(projects_df)):
            y = 0
            while (
                y + int(projects_df.iloc[x, projects_df.columns.get_loc("level_of_effort")])
                < height_of_matrix
            ):
                project_dates_and_effort_df = project_plotting_df.loc[
                    str(
                        y
                        + int(
                            projects_df.iloc[
                                x, projects_df.columns.get_loc("level_of_effort")
                            ]
                        )
                        - 1
                    ) : str(y),
                    projects_df.iloc[
                        x, projects_df.columns.get_loc("start_date")
                    ] : projects_df.iloc[x, projects_df.columns.get_loc("end_date")],
                ]

                if project_dates_and_effort_df.equals(
                    master_plotting_df.loc[
                        str(
                            y
                            + int(
                                projects_df.iloc[
                                    x, projects_df.columns.get_loc("level_of_effort")
                                ]
                            )
                            - 1
                        ) : str(y),
                        projects_df.iloc[
                            x, projects_df.columns.get_loc("start_date")
                        ] : projects_df.iloc[x, projects_df.columns.get_loc("end_date")],
                    ]
                ):
                    projects_df.iloc[x, projects_df.columns.get_loc("stack")] = y

                    master_plotting_df.loc[
                        str(
                            y
                            + int(
                                projects_df.iloc[
                                    x, projects_df.columns.get_loc("level_of_effort")
                                ]
                            )
                            - 1
                        ) : str(y),
                        projects_df.iloc[
                            x, projects_df.columns.get_loc("start_date")
                        ] : projects_df.iloc[x, projects_df.columns.get_loc("end_date")],
                    ] = master_plotting_df.loc[
                        str(
                            y
                            + int(
                                projects_df.iloc[
                                    x, projects_df.columns.get_loc("level_of_effort")
                                ]
                            )
                            - 1
                        ) : str(y),
                        projects_df.iloc[
                            x, projects_df.columns.get_loc("start_date")
                        ] : projects_df.iloc[x, projects_df.columns.get_loc("end_date")],
                    ].applymap(
                        lambda z: 1
                    )

                    y += 1
                    break

                else:
                    y += 1

        new_max_height_df = projects_df.copy()
        new_max_height = projects_df["stack"].max()
        new_max_height_df = projects_df[projects_df["stack"] == new_max_height]
        new_max_height_plus_level_of_effort = (
            int(new_max_height_df["level_of_effort"].max()) + new_max_height
        )

        # Sort by 'Start date' (ascending), 'Due date' (descending), and 'level_of_effort' (ascending)

        df = projects_df.copy()

        fig, gnt = plt.subplots(figsize=(16, 10))
        array = np.linspace(0, 1, len(df))
        np.random.shuffle(array)
        color = iter(cm.rainbow(array))

        df = df.reset_index()
        for l in range(0, len(df)):
            start = df.loc[l, "start_date"]
            finish = df.loc[l, "end_date"]

            



    # Use the wrap_text function to wrap the Title field for the label
            gnt.broken_barh(
                [(pd.to_datetime(start), pd.to_datetime(finish) - pd.to_datetime(start))],
                [int(df.loc[l, "stack"]), int(df.loc[l, "level_of_effort"])],
                color=next(color),
                label=wrap_text(df.loc[l, "Title"])
            )
            # gnt.broken_barh(
            #     [(pd.to_datetime(start), pd.to_datetime(finish) - pd.to_datetime(start))],
            #     [int(df.loc[l, "stack"]), int(df.loc[l, "level_of_effort"])],
            #     color=next(color),
            #     label=df.loc[l, "Title"],
            # )

            data = [(pd.to_datetime(start), pd.to_datetime(finish) - pd.to_datetime(start))]

            for x1, x2 in data:
                gnt.text(
                    x=x1 + x2 / 2,
                    y=(int(df.loc[l, "stack"]) + int(df.loc[l, "level_of_effort"]))
                    - int(df.loc[l, "level_of_effort"]) / 2,
                    s=df.loc[l, "Title"],
                    ha="center",
                    va="center",
                    color="blue",
                    fontsize="small",
                )

        fig.tight_layout()
        gnt.set_xlabel("Date")
        gnt.set_ylabel("Effort Level")
        fig.legend(ncol=3)

        #top_value_benchmark = 0.710 / 10
        #top_value = top_value_benchmark * new_max_height_plus_level_of_effort

        #plt.subplots_adjust(left=0.1, right=0.9, bottom=0.2, top=0.9)
        #plt.xticks(rotation=45)
        # plt.show(block=True)
        # Generate the plot
        img = BytesIO()
        plt.savefig(img, format="png")
        img.seek(0)

        
        return img
        #plot_url = base64.b64encode(img.getvalue()).decode("utf8")

        # Use send_file to return the image for download
        #return send_file(img, mimetype='image/png', as_attachment=True, download_name='chart.png')
        # Create static/images directory if it doesn't exist
        #os.makedirs('static/images', exist_ok=True)

        # Save the file
        #plt.savefig('static/images/chart.png')
    except Exception as e:
        print(f"Error generating Gantt chart: {e}")
        raise