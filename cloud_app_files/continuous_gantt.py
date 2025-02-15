
import pandas as pd
import plotly.express as pex
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from datetime import datetime
from matplotlib.pyplot import cm
import numpy as np
from io import BytesIO
import os

def generate_gantt_chart():
    uploaded_files = os.listdir("./uploads")
    if not uploaded_files:
        raise Exception("No uploaded file found")

    # Assuming we are working with the most recently uploaded file.
    latest_file = uploaded_files[-1]
    filename = os.path.join("./uploads", latest_file)
    
    wb = load_workbook(filename)
    ws = wb.active

    rows = [[c.value for c in r] for r in ws.iter_rows() if not ws.row_dimensions[r[0].row].hidden]
    projects_df = pd.DataFrame(data=rows[1:], columns=rows[0], dtype=str)

    projects_df = projects_df.dropna(subset=["Start date", "Due date"])
    projects_df["start_date"] = projects_df["start_date"].apply(lambda x: x.replace(" 00:00:00", ""))
    projects_df["end_date"] = projects_df["end_date"].apply(lambda x: x.replace(" 00:00:00", ""))

    df = projects_df.copy()
    fig, gnt = plt.subplots(figsize=(16, 10))
    array = np.linspace(0, 1, len(df))
    np.random.shuffle(array)
    color = iter(cm.rainbow(array))

    for l in range(len(df)):
        start = datetime.strptime(df.loc[l, "start_date"], "%Y-%m-%d")
        finish = datetime.strptime(df.loc[l, "end_date"], "%Y-%m-%d")
        gnt.broken_barh([(pd.to_datetime(start), pd.to_datetime(finish) - pd.to_datetime(start))],
                        [int(df.loc[l, "stack"]), int(df.loc[l, "level_of_effort"])],
                        color=next(color),
                        label=df.loc[l, "task"])

    fig.tight_layout()
    img = BytesIO()
    plt.savefig(img, format="png")
    img.seek(0)

    return img
