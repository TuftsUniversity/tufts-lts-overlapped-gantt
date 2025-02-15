import pandas as pd
import plotly.express as pex
import matplotlib.pyplot as plt
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from datetime import datetime
from matplotlib.pyplot import cm
import numpy as np
import os

@app.route('/generate', methods=['POST'])
def generate_gantt_chart():

    uploaded_files = os.listdir(app.config['UPLOAD_FOLDER'])

    if not uploaded_files:
        return jsonify({'message': 'No uploaded file found'}), 400

    # Assuming we are working with the most recently uploaded file.
    latest_file = uploaded_files[-1]  # Modify as needed if multiple files

    filename = os.path.join(app.config['UPLOAD_FOLDER'], latest_file)

    #filename = askopenfilename(title="Select project sheet")
    wb = load_workbook(filename)
    ws = wb.active

    rows = [
        [c.value for c in r]
        for r in ws.iter_rows()
        if not ws.row_dimensions[r[0].row].hidden
    ]

    projects_df = pd.DataFrame(data=rows[1:], columns=rows[0], dtype=str)
    #projects_df["level_of_effort"] = projects_df["level_of_effort"].astype("float")
    #projects_df["level_of_effort"] = projects_df["level_of_effort"].astype("Int64")

    projects_df["level_of_effort"] = 2
    projects_df = projects_df[["Issue", "level_of_effort", "Start date", "Due date"]].copy()
    projects_df["stack"] = 0
    projects_df = projects_df.sort_values(["Start date", "Due date", "level_of_effort"])
    pd.set_option("display.max_columns", None)
    pd.options.display.max_colwidth = 200

    projects_df = projects_df.dropna(subset=['Start date', 'Due date'])
    projects_df["start_date"] = projects_df["start_date"].apply(
        lambda x: x.replace(" 00:00:00", "")
    )
    projects_df["end_date"] = projects_df["end_date"].apply(
        lambda x: x.replace(" 00:00:00", "")
    )

    min_start_date = (
        projects_df["start_date"].apply(lambda x: datetime.strptime(x, "%M/%D/%Y")).min()
    )

    max_end_date = (
        projects_df["end_date"].apply(lambda x: datetime.strptime(x, "%M/%D/%Y")).max()
    )
    delta = max_end_date - min_start_date
    length_of_matrix = delta

    delta = int(delta.total_seconds() / 60 / 60 / 24)
    height_of_matrix = int(projects_df["level_of_effort"].sum())

    rows, cols = (delta, height_of_matrix)
    arr = [[0] * cols] * rows

    date_range = pd.date_range(min_start_date, max_end_date)

    range_list = list(reversed(list(range(0, height_of_matrix))))

    for z in range(0, len(range_list)):
        range_list[z] = str(range_list[z])

    master_plotting_df = pd.DataFrame(columns=date_range, index=range_list)
    master_plotting_df = master_plotting_df.applymap(lambda x: 0)

    project_plotting_df = master_plotting_df.copy()
    for x in range(0, len(projects_df)):
        y = 0
        while (
            y + int(projects_df.iloc[x, projects_df.columns.get_loc("level_of_effort")])
            < height_of_matrix
        ):
            project_dates_and_effort_df = project_plotting_df.loc[
                str(
                    y
                    + int(
                        projects_df.iloc[x, projects_df.columns.get_loc("level_of_effort")]
                    )
                    - 1
                ) : str(y),
                projects_df.iloc[
                    x, projects_df.columns.get_loc("start_date")
                ] : projects_df.iloc[x, projects_df.columns.get_loc("end_date")],
            ]

            if project_dates_and_effort_df.equals(
                master_plotting_df.loc[
                    str(
                        y
                        + int(
                            projects_df.iloc[
                                x, projects_df.columns.get_loc("level_of_effort")
                            ]
                        )
                        - 1
                    ) : str(y),
                    projects_df.iloc[
                        x, projects_df.columns.get_loc("start_date")
                    ] : projects_df.iloc[x, projects_df.columns.get_loc("end_date")],
                ]
            ):
                projects_df.iloc[x, projects_df.columns.get_loc("stack")] = y

                master_plotting_df.loc[
                    str(
                        y
                        + int(
                            projects_df.iloc[
                                x, projects_df.columns.get_loc("level_of_effort")
                            ]
                        )
                        - 1
                    ) : str(y),
                    projects_df.iloc[
                        x, projects_df.columns.get_loc("start_date")
                    ] : projects_df.iloc[x, projects_df.columns.get_loc("end_date")],
                ] = master_plotting_df.loc[
                    str(
                        y
                        + int(
                            projects_df.iloc[
                                x, projects_df.columns.get_loc("level_of_effort")
                            ]
                        )
                        - 1
                    ) : str(y),
                    projects_df.iloc[
                        x, projects_df.columns.get_loc("start_date")
                    ] : projects_df.iloc[x, projects_df.columns.get_loc("end_date")],
                ].applymap(
                    lambda z: 1
                )

                y += 1
                break

            else:
                y += 1

    new_max_height_df = projects_df.copy()
    new_max_height = projects_df["stack"].max()
    new_max_height_df = projects_df[projects_df["stack"] == new_max_height]
    new_max_height_plus_level_of_effort = (
        int(new_max_height_df["level_of_effort"].max()) + new_max_height
    )

    df = projects_df.copy()

    fig, gnt = plt.subplots(figsize=(16, 10))
    array = np.linspace(0, 1, len(df))
    np.random.shuffle(array)
    color = iter(cm.rainbow(array))

    df = df.reset_index()
    for l in range(0, len(df)):
        start = datetime.strptime(df.loc[l, "start_date"], "%Y-%m-%d")
        finish = datetime.strptime(df.loc[l, "end_date"], "%Y-%m-%d")
        gnt.broken_barh(
            [(pd.to_datetime(start), pd.to_datetime(finish) - pd.to_datetime(start))],
            [int(df.loc[l, "stack"]), int(df.loc[l, "level_of_effort"])],
            color=next(color),
            label=df.loc[l, "task"],
        )

        data = [(pd.to_datetime(start), pd.to_datetime(finish) - pd.to_datetime(start))]

        for x1, x2 in data:
            gnt.text(
                x=x1 + x2 / 2,
                y=(int(df.loc[l, "stack"]) + int(df.loc[l, "level_of_effort"]))
                - int(df.loc[l, "level_of_effort"]) / 2,
                s=df.loc[l, "task"],
                ha="center",
                va="center",
                color="blue",
                fontsize="xx-small",
            )

    fig.tight_layout()
    gnt.set_xlabel("Date")
    gnt.set_ylabel("Hours per Day")
    fig.legend(ncol=3)

    top_value_benchmark = 0.710 / 10
    top_value = top_value_benchmark * new_max_height_plus_level_of_effort

    plt.subplots_adjust(left=0.1, right=0.9, bottom=0.2, top=0.9)
    plt.xticks(rotation=45)
    #plt.show(block=True)
    # Generate the plot
    img = BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode('utf8')

    return render_template('index.html', plot_url=plot_url)
